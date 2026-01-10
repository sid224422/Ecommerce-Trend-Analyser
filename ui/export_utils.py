"""
Export utilities for Market Analytics System

Handles exporting analysis results in various formats (JSON, CSV, Excel, PDF).
"""

import json
import pandas as pd
from pathlib import Path
from datetime import datetime
from typing import Dict, Optional
from io import BytesIO

try:
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import Font, Alignment, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib import colors
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False


def export_to_json(results: Dict, include_llm: bool = True) -> str:
    """Export results to JSON format."""
    export_data = {
        "timestamp": results.get("timestamp", datetime.now().isoformat()),
        "total_records": results.get("total_records", 0),
        "agent_outputs": [
            results["agents"]["brand"],
            results["agents"]["pricing"],
            results["agents"]["feature"],
            results["agents"]["gap"]
        ]
    }
    
    if include_llm and results.get('llm_summary'):
        export_data["llm_summary"] = results['llm_summary']
    else:
        export_data["llm_summary"] = None
    
    return json.dumps(export_data, indent=2, ensure_ascii=False)


def export_to_csv(results: Dict) -> str:
    """Export results to CSV format (creates summary CSV)."""
    data_rows = []
    
    # Brand data
    for brand in results["agents"]["brand"]["results"]["top_brands"]:
        data_rows.append({
            "Category": "Brand",
            "Name": brand["brand"],
            "Count": brand["count"],
            "Confidence": brand["confidence"],
            "Details": ""
        })
    
    # Feature data
    for feature in results["agents"]["feature"]["results"]["top_features"]:
        data_rows.append({
            "Category": "Feature",
            "Name": feature["feature"],
            "Count": feature["count"],
            "Confidence": feature["confidence"],
            "Details": ""
        })
    
    # Gap data
    for gap in results["agents"]["gap"]["results"]["top_gaps"]:
        data_rows.append({
            "Category": "Gap",
            "Name": f"{gap['brand']} - {gap['feature']}",
            "Count": gap["observed_count"],
            "Confidence": gap["gap_score"],
            "Details": f"Expected: {gap['expected_count']:.1f}"
        })
    
    df = pd.DataFrame(data_rows)
    return df.to_csv(index=False)


def export_to_excel(results: Dict, include_charts: bool = False) -> bytes:
    """Export results to Excel format."""
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
    
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Sheet 1: Summary
    ws_summary = wb.create_sheet("Summary", 0)
    ws_summary.append(["Market Analytics Report"])
    ws_summary.append(["Generated", results.get("timestamp", datetime.now().isoformat())])
    ws_summary.append(["Total Records", results.get("total_records", 0)])
    ws_summary.append([])
    
    # Sheet 2: Brands
    ws_brands = wb.create_sheet("Brands", 1)
    brands_df = pd.DataFrame(results["agents"]["brand"]["results"]["top_brands"])
    for r in dataframe_to_rows(brands_df, index=False, header=True):
        ws_brands.append(r)
    
    # Format header
    for cell in ws_brands[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Sheet 3: Features
    ws_features = wb.create_sheet("Features", 2)
    features_df = pd.DataFrame(results["agents"]["feature"]["results"]["top_features"])
    for r in dataframe_to_rows(features_df, index=False, header=True):
        ws_features.append(r)
    
    # Format header
    for cell in ws_features[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Sheet 4: Pricing
    ws_pricing = wb.create_sheet("Pricing", 3)
    pricing_stats = results["agents"]["pricing"]["results"]["price_statistics"]
    ws_pricing.append(["Metric", "Value"])
    for key, value in pricing_stats.items():
        ws_pricing.append([key.replace("_", " ").title(), value])
    
    # Format header
    for cell in ws_pricing[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Sheet 5: Gaps
    if results["agents"]["gap"]["results"]["top_gaps"]:
        ws_gaps = wb.create_sheet("Market Gaps", 4)
        gaps_df = pd.DataFrame(results["agents"]["gap"]["results"]["top_gaps"])
        for r in dataframe_to_rows(gaps_df, index=False, header=True):
            ws_gaps.append(r)
        
        # Format header
        for cell in ws_gaps[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Save to bytes
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def export_to_pdf(results: Dict, include_charts: bool = False) -> bytes:
    """Export results to PDF format."""
    if not REPORTLAB_AVAILABLE:
        raise ImportError("reportlab is required for PDF export. Install with: pip install reportlab")
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1f77b4'),
        spaceAfter=30,
        alignment=1  # Center
    )
    story.append(Paragraph("Market Analytics Report", title_style))
    story.append(Spacer(1, 12))
    
    # Summary
    story.append(Paragraph(f"Generated: {results.get('timestamp', datetime.now().isoformat())}", styles['Normal']))
    story.append(Paragraph(f"Total Records Analyzed: {results.get('total_records', 0):,}", styles['Normal']))
    story.append(Spacer(1, 20))
    
    # Brand Analysis
    story.append(Paragraph("Brand Analysis", styles['Heading2']))
    brand_data = [["Brand", "Count", "Market Share"]]
    for brand in results["agents"]["brand"]["results"]["top_brands"][:10]:
        brand_data.append([
            brand["brand"],
            str(brand["count"]),
            f"{brand['confidence']*100:.2f}%"
        ])
    
    brand_table = Table(brand_data)
    brand_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(brand_table)
    story.append(Spacer(1, 20))
    
    # Pricing Analysis
    story.append(Paragraph("Pricing Analysis", styles['Heading2']))
    pricing_stats = results["agents"]["pricing"]["results"]["price_statistics"]
    pricing_data = [
        ["Metric", "Value"],
        ["Min Price", f"${pricing_stats['min_price']:.2f}"],
        ["Max Price", f"${pricing_stats['max_price']:.2f}"],
        ["Mean Price", f"${pricing_stats['mean_price']:.2f}"],
        ["Median Price", f"${pricing_stats['median_price']:.2f}"]
    ]
    
    pricing_table = Table(pricing_data)
    pricing_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(pricing_table)
    story.append(Spacer(1, 20))
    
    # Feature Analysis
    story.append(Paragraph("Feature Analysis", styles['Heading2']))
    feature_data = [["Feature", "Count", "Prevalence"]]
    for feature in results["agents"]["feature"]["results"]["top_features"][:10]:
        feature_data.append([
            feature["feature"],
            str(feature["count"]),
            f"{feature['confidence']*100:.2f}%"
        ])
    
    feature_table = Table(feature_data)
    feature_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(feature_table)
    story.append(Spacer(1, 20))
    
    # Gap Analysis
    if results["agents"]["gap"]["results"]["top_gaps"]:
        story.append(Paragraph("Market Gap Analysis", styles['Heading2']))
        gap_data = [["Brand", "Feature", "Gap Score", "Opportunity"]]
        for gap in results["agents"]["gap"]["results"]["top_gaps"][:10]:
            gap_data.append([
                gap["brand"],
                gap["feature"][:30] + "..." if len(gap["feature"]) > 30 else gap["feature"],
                f"{gap['gap_score']:.3f}",
                f"{int(gap['expected_count'] - gap['observed_count'])} products"
            ])
        
        gap_table = Table(gap_data)
        gap_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(gap_table)
    
    # LLM Summary if available
    if results.get('llm_summary') and results['llm_summary'].get('status') == 'success':
        story.append(PageBreak())
        story.append(Paragraph("AI Summary", styles['Heading2']))
        story.append(Paragraph(results['llm_summary']['summary'], styles['Normal']))
    
    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()
